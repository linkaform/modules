# -*- coding: utf-8 -*-
import pyexcel, openpyxl, os, requests
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
from datetime import datetime, timedelta, date, time
from pytz import timezone

class PCI_Utils():
    def __init__(self, cr=None, cr_admin=None, lkf_api=None, net=None, settings=None, lkf_obj=None):
        self.cr = cr
        self.cr_admin = cr_admin
        self.lkf_api = lkf_api
        self.net = net
        self.settings = settings
        self.lkf_obj = lkf_obj
        
        self.list_tareas_solo_vsi = ['TS1L7VGPI', 'TS2L7VGPI']
        
        self.FORM_ID_PLANTILLA_EMPRESARIAL = 46429
        self.FORM_ID_EXP_TECNICOS = 63122
        self.FORM_ID_PSR = 123416
        self.FORM_ID_EMAIL_ERRORES = 120994
        self.FORM_ID_AUTORIZA_CARGA_FOLIOS = 62694
        self.FORM_ID_CAMBIO_TECNOLOGIA = 120499
        self.FORM_ID_HISTORICO_DE_PAGOS = 19704
        self.FORM_ID_LISTADO_TECNICOS = 12165
        self.UPLOAD_PERMISIONS = 21472
        
        self.CATALOG_ID_CONTRATISTAS_1_0_ADMIN = 59273
        self.CATALOG_ID_COPES = 46944
        self.CATALOG_ID_TIPOS_TAREA_FACTIBLES = 56269

        self.equivalcens_map = { 
            'Aerea': ['AEREA', 'M. AEREO'],
            'Alfanumérico':['SERIE ONT ALFANUMÉRICO','ONT ALFANUMÉRICO','SERIE ONT ALFANUMERICO','Serie ONT Alfanumerico', 'Serie ONT Alfanumérico',
                'Serie ONT alfanumerico','Serie ONT alfanumérico','serie ONT alfanumerico', 'serie ONT alfanumérico','serie ONT Alfanumerico', 'serie ONT Alfanumérico'],
            'Clase de Servicio': ['CLASE_SERV', 'Clase', 'Clase Servicio', 'Clase de Servicio'],
            'Created At': ['Fecha Pendiente', 'Fecha de Pendiente', 'FECHA DE PENDIENTES'],
            'Contratista': ['CONTRATISTA'],
            'Dilacion': ['Dilación'],
            'Estatus de Orden':['ESTATUS de Orden', 'Estatus de Orden','Estatus','ESTATUS'],
            'Etapa': ['Etapa'],
            'Expediente del Tecnico': ['EXP. TEC.', 'EXPEDIENTE', 'EXPEDIENTE TÉCNICO', 'EXPEDIENTE TECNICO', 'EXPEDIENTE DEL TÉCNICO', 'EXPEDIENTE DEL TECNICO', 
                'Expediente del Técnico', 'Expediente del Tecnico', 'Expediente Técnico', 'Expediente técnico', 'Expediente tecnico', 'Expediente', 'expediente'],
            'Expediente Del Tecnico': ['EXP. TEC.', 'EXPEDIENTE', 'EXPEDIENTE TÉCNICO', 'EXPEDIENTE TECNICO', 'EXPEDIENTE DEL TÉCNICO', 'EXPEDIENTE DEL TECNICO', 
                'Expediente del Técnico', 'Expediente del Tecnico', 'Expediente Técnico', 'Expediente técnico', 'Expediente tecnico', 'Expediente', 'expediente'],
            'Fecha Contratada': ['F_CONTRATA'],
            'Folio': ['Folio', 'Folio Pisa', 'FolioPisa', 'FOLIO', 'Orden de Servicio', 'ORDEN DE SERVICIO','O.S', 'OS'],
            'Fol_Pisaplex': ['PISAPLEX'],
            'Metros Bajante':['MTS. BAJANTE (PARALELO)'],
            'ONT Numerico': ['SERIE ONT NUMERICO', 'SERIE ONT NUMÉRICO', 'Serie ONT numerico', 'Serie ONT numérico', 'Serie ONT Numerico', 'Serie ONT Numérico',
                'serie ONT numerico', 'serie ONT numérico', 'serie ONT Numerico', 'serie ONT Numérico', 'ONT NUMERICO', 'ONT NUMÉRICO', 'ONT Numérico', 'ont_numerico', 'ont_numérico'],
            'Subterranea': ['SUBTERRANEA','SUBTERRANEO', 'M. SUBTERRANEO'],
            'Teléfono':['TELEFONO (DIEZ DIGITOS)','TELEFONO','TELEFONO ( A 10 DIGITOS)'],
            'Telefono':['Telefono','Teléfono'],
            'Tecnico': ['TECNICO','nombre_tecnico','TÉCNICO','técnico','Técnico','Tecnico','tecnico'],
            'Terminal Optica':['TERMINAL OPTICA ', 'TERMINAL', 'TERMINAL ', 'Terminal ', 'TERMINAL ÓPTICA ', 'TERMINAL ÓPTICA'],
            'Tipo':['TIPO', 'TIPO DE OS', 'TIPO TAREA', 'TIPO DE O.S.', 'Tipo de Tarea'],
            'Tipo de Tarea': ['TIPO TAREA', 'TIPO DE OS', 'TIPO DE O.S.'],
            'Fecha de Liquidacion': ['Fecha Liquidada','Fecha De Liquidacion','Fecha de Liquidacion','FECHA DE LIQUIDACION','fecha de liquidacion',
                'Fecha De Liquidación','Fecha de Liquidación','FECHA DE LIQUIDACIÓN','fecha de liquidación'],
            'Fecha de Asignacion': ['Fecha De Asignacion','Fecha de Asignacion','FECHA DE ASIGNACION','fecha de asignacion',
                'Fecha De Asignación','Fecha de Asignación','FECHA DE ASIGNACIÓN','fecha de asignación'],
            'Fecha Liquidada': ['Fecha de Liquidacion'],
            'Cope': ['Cope','COPE','cope'],
            'Motivo de Objecion': ['DETALLE'],
            'usuario_reporta': ['usuario_reporta'],
            'Num. Serie': ['Num. Serie', 'Num Serie'],
            'cambio_tecnologia': ['cambio_tecnologia', 'cambio_tecnología', 'cambio_de_tecnologia', 'cambio_de_tecnología'],
        }

    def add_coma(self, snum):
        return "{:,.2f}".format( float(snum) )

    def adjust_metros_fibra(self, metros):
        #regresa el metraje segun seleccion de orden de servico fibra
        metros = int(metros)

        if metros > 300:
            return '300'

        mod = metros%25
        if metros <= 200:
            return str(metros - mod + 25) if mod > 0 else metros
        elif metros <= 300:
            if mod >0:
                return '300' if metros >250 else '250'
        return str(metros)

    def arregla_msg_error_sistema(self, response):
        try:
            msg_err_arr = []
            data_str_err = ''
            if response.get('status_code', 0) == 400:
                data_str_err = "Formato incorrecto:"
            for msg_fin in response.get('json',{}):
                info_json = response['json'][msg_fin]
                msgs = info_json.get('msg', [])
                if msgs:
                    msg_err = str(msgs[0])
                    label_err = info_json.get('label')
                    msg_err_arr.append(str(msg_err+':'+label_err))
                else:
                    for i_err in info_json:
                        info_i = info_json[i_err]
                        for id_group in info_i:
                            info_group = info_i[id_group]
                            msg_err = str(info_group['msg'][0])
                            label_err = info_group['label']
                            msg_err_arr.append('SET {0}:: {1} - {2}'.format(i_err, msg_err, label_err))
            if msg_err_arr:
                data_str_err = list_to_str(msg_err_arr)
        except Exception as e:
            print('response', response)
            print('Exception =  ', e)
            data_str_err = "Ocurrió un error desconocido favor de contactar a soporte"
            error_json = response.get('json',{}).get('error','')
            if error_json:
                data_str_err = error_json
        return data_str_err

    def asignar_registro_a_conexion(self, connection_id, record_assign):
        id_record_assign = f'/api/infosync/form_answer/{record_assign}/'
        return self.lkf_api.assigne_connection_records( connection_id, [id_record_assign,])

    def find_in_lista_tecnicos(self, id_contratista, expediente_telmex, find_by_expediente=False, find_in_admin=False):
        query = {
            'form_id': self.lkf_obj.FORM_ID_EXPEDIENTES_DE_TECNICOS, 'deleted_at': {'$exists':False},
            'answers.590a4761f851c20e60ac168c': 'activo',
            'answers.f1216500a010000000000001': int(expediente_telmex)
        }
        obj_id_contratistas_1_0 = '5f344a0476c82e1bebc991d5' if find_in_admin else self.lkf_obj.CATALOGO_CONTRATISTAS_OBJ_ID
        if not find_by_expediente:
            query.update({
                'answers.{}.5f344a0476c82e1bebc991d6'.format(obj_id_contratistas_1_0): {'$in': [ str(id_contratista) ]},
            })
        # print('-----query Listado:',query)
        project_query = {
            '$project': {
                'id_usuario': {'$arrayElemAt': ['$answers.{}.5f344a0476c82e1bebc991d6'.format(obj_id_contratistas_1_0),0]},
                'expediente': '$answers.f1216500a010000000000001'
            }
        }
        if find_in_admin:
            query.update({'form_id': self.FORM_ID_EXP_TECNICOS})
            print('\n+++ query listado tecnicos find_in_admin =',query)
            records = self.cr_admin.aggregate([{'$match': query}, project_query])
        else:
            print('\n+++ query listado tecnicos =',query)
            records = self.cr.aggregate([{'$match': query}, project_query])
        return {
            int(rec_exp.get('id_usuario')): {'expediente': rec_exp['expediente']} for rec_exp in records if rec_exp.get('id_usuario')
        }
    
    def create_record_expediente(self, answers_to_expediente, properties_expediente, form_id_expediente, jwt_settings_key=False):
        metadata_expediente = self.lkf_api.get_metadata( form_id_expediente )
        metadata_expediente['properties'] = properties_expediente
        metadata_expediente['answers'] = answers_to_expediente
        resp_create_exp = self.lkf_api.post_forms_answers(metadata_expediente, jwt_settings_key=jwt_settings_key)
        print('resp_create_exp =',resp_create_exp)
        return resp_create_exp

    def calcula_numero_semana(self):
        hoy = datetime.today()
        number_week = hoy.isocalendar()[1]
        week_day = hoy.weekday()
        if week_day == 6:
            number_week += 1
        if hoy.month == 1 and number_week > 50:
            number_week = 1
        return number_week

    def carga_prod_find_expediente(self, connection_id, parent_connection, numero_expediente, **kwargs):
        # Se busca el expediente para ver si le pertenece al contratista o no
        # Hay ocasiones muy raras en que existe mas de un registro de expediente, se prepara lista de expedientes y contratistas
        
        # connections_expediente = self.get_expedientes_contratistas(list_expedientes=[int(numero_expediente)], with_item_idUser=True)
        connections_expediente = self.find_in_lista_tecnicos(None, numero_expediente, find_by_expediente=True, find_in_admin=True)
        found_in_admin = parent_connection in connections_expediente

        connections_expediente_iasa = self.find_in_lista_tecnicos(None, numero_expediente, find_by_expediente=True)
        found_in_iasa = connection_id in connections_expediente_iasa

        # Caso ideal, si el expediente le pertenece a IASA y al contratista que hace la carga
        if found_in_admin and found_in_iasa:
            return True

        # Caso donde si existe el expediente ya sea en Admin o en IASA lo dejo avanzar para que lo detenga la otra validacion y marque error donde haga falta
        if connections_expediente or connections_expediente_iasa:
            return False

        # Si no existe el expediente en Admin ni en IASA pues se debe crear
        if not connections_expediente and not connections_expediente_iasa:
            print('expediente {} no encontrado... creando'.format(numero_expediente))

            def get_data_to_connection(dict_data, conexion):
                return { # Catalogo de Contratistas
                    '5f344a0476c82e1bebc991d7': dict_data.get('nombre_contratista'),
                    '5f344a0476c82e1bebc991db': [dict_data.get('razon_social')],
                    '5f344a0476c82e1bebc991da': [dict_data.get('division_contratista')],
                    '5f344a0476c82e1bebc991d9': [dict_data.get('telefono_contratista')],
                    '5f344a0476c82e1bebc991d8': [dict_data.get('correo_contratista')],
                    '5f344a0476c82e1bebc991d6': [str(conexion)],
                    '673e4bc4f3e1a7191de583cf': [dict_data.get('socio_comercial')],
                }

            info_conn_iasa = kwargs.get('info_iasa', {})
            answers_to_expediente = {
                '5d77be2577248fa5ad7f82ee': { # CATALOGO DE COPES 
                    '5d641731ddd3adcc24778a9d': kwargs.get('cope'),
                    '5d641731ddd3adcc24778a9e': [kwargs.get('abreviatura_cope')],
                    '5d641731ddd3adcc24778a9c': [kwargs.get('area')],
                    '5d641731ddd3adcc24778a9b': [kwargs.get('subdireccion')],
                    '5d77bd96eeb32658ce522ec4': [kwargs.get('division_pic_movil')],
                    '5d641731ddd3adcc24778a9a': [kwargs.get('division')]
                },
                '5fa470ac2bf44d032891457e': kwargs.get('nombre_tecnico'),
                'f1216500a010000000000001': int(numero_expediente),
                '590a4761f851c20e60ac168c': 'activo',
                '58db0090b43fdd5c1419dae2': 'tecnico',
                '5f344a0476c82e1bebc991d5': get_data_to_connection(info_conn_iasa, 1940) # Catalogo de Contratistas
            }
            properties_expediente = {"device_properties":{"system": "SCRIPT","process":"CARGA DE PRODUCCION", "accion":'Crear Expediente', "folio carga":kwargs.get('current_record_folio'), "archive":"iasa_carga_produccion_hibrido.py"}}
            
            print('--- --- --- Creando registro de Expediente en la cuenta de Admin')

            # OJO AQUI. LOS DATOS DEL CONTRATISTA DEBEN SER LOS DE IASA

            resp_exp_admin = self.create_record_expediente(answers_to_expediente, properties_expediente, self.FORM_ID_EXP_TECNICOS)
            if resp_exp_admin.get('status_code') == 201:
                self.send_notification_email('Nuevo Expediente desde Carga de Produccion', 'Se creó un nuevo expediente: {} para el contratista: {} desde la carga: {}'.format(numero_expediente, info_conn_iasa.get('nombre_contratista'), kwargs.get('current_record_folio')), properties_expediente)
                
                print('--- --- --- Creando registro de Expediente en la cuenta de IASA')

                # OJO AQUI. LOS DATOS DEBEN SER LOS DEL CONTRATISTA DE IASA Y APUNTAR A LA BD DE IASA

                answers_to_expediente[self.lkf_obj.CATALOGO_CONTRATISTAS_OBJ_ID] = get_data_to_connection(kwargs, connection_id) # Catalogo de Contratistas
                answers_to_expediente.pop('5f344a0476c82e1bebc991d5', None)

                resp_exp_iasa = self.create_record_expediente(answers_to_expediente, properties_expediente, self.lkf_obj.FORM_ID_EXPEDIENTES_DE_TECNICOS, jwt_settings_key='JWT_KEY_IASA')

                # El expediente se creó correctamente en las cuentas de Admin y IASA
                if resp_exp_iasa.get('status_code') == 201:
                    return True
            
        return False

    def catch_repeted_keys(self, existing_keys, actual_key):
        if actual_key in existing_keys:
            next_key = 0
            for key in existing_keys:
                if key == actual_key:
                    next_key = 1
                if str(key).find('-') > 0:
                    if key.split('-')[0] == actual_key:
                        if int(key.split('-')[1]) + 1 > next_key:
                            next_key = key.split('-')[1] + 1
            if next_key > 0 :
                return str(actual_key) + '-' + str(next_key)
        return actual_key

    def check_folio(self, form_id_admin, folio, telefono, area):
        # form_id_admin = self.dict_equivalences_forms_id[form_id]
        query_folio_os = {
            'form_id':  form_id_admin, 'deleted_at' : {'$exists':False}, 
            'folio': folio,
            'answers.f1054000a010000000000005': telefono,
            'answers.f1054000a0100000000000a2': area
        }
        print('query_folio_os =',query_folio_os)
        record = self.cr_admin.find_one(query_folio_os, {'folio':1, 'answers':1, 'connection_id':1, 'user_id':1, 'created_at': 1})
        return record

    def check_folio_pagado(self, folio):
        record = self.cr_admin.find_one({
            'form_id':  self.FORM_ID_HISTORICO_DE_PAGOS, 'deleted_at' : {'$exists':False},
            'folio': folio, 'answers.f138450000000000000000f2':'pagado'
        }, {'folio':1, 'answers':1})
        return record

    def compara_fecha_rango(self, fecha_a_validar):
        if isinstance(fecha_a_validar, str):
            fecha_a_validar = datetime.strptime(fecha_a_validar, '%d/%m/%Y')
        fecha_hoy = date.today()
        fecha_pasada = date(2019, 1, 1)
        fecha_hoy = datetime.combine(fecha_hoy, time(0, 0))
        fecha_pasada = datetime.combine(fecha_pasada, time(0, 0))
        fecha_a_validar = datetime.combine(fecha_a_validar, time(0, 0))

        if fecha_hoy < fecha_a_validar or fecha_a_validar < fecha_pasada:
            return False
        else:
            return True

    def compartir_forma(self, user_id, form_id, permissions="can_share_item", jwt_settings_key=None):
        objects = {
            "file_shared" :  f"/api/infosync/item/{form_id}/",
            "owner": f"/api/infosync/user/{user_id}/",
            "perm": permissions
        }
        return self.lkf_api.share_form(objects, jwt_settings_key=jwt_settings_key)

    def complementos_cuentas_contratistas(self, cuentas_contratistas, autorizacion_facturar):
        for cuenta_contratista in cuentas_contratistas:
            contratista = cuenta_contratista.get(self.lkf_obj.CATALOGO_CONTRATISTAS_OBJ_ID, {}).get('5f344a0476c82e1bebc991d7', '')
            mango_query = {"selector": {"answers": {"$and":[ {'5f344a0476c82e1bebc991d7': {'$eq': contratista}} ]} }, "limit": 1, "skip": 0}
            row_catalog = self.lkf_api.search_catalog( self.lkf_obj.CATALOGO_CONTRATISTAS_ID, mango_query )
            if row_catalog:
                info_row = row_catalog[0]

                print('+++ data to update= ', {'619e7a46c79af2f6eaf888c5': autorizacion_facturar} )
                print('+++ id catalog= ', self.lkf_obj.CATALOGO_CONTRATISTAS_ID )
                print('+++ record_id= ', info_row[ '_id' ])

                res_update = self.lkf_api.update_catalog_multi_record({'619e7a46c79af2f6eaf888c5': autorizacion_facturar}, self.lkf_obj.CATALOGO_CONTRATISTAS_ID, record_id=[ info_row[ '_id' ], ])
                print('res_update =',res_update)

    def delete_record(self, id_record='', jwt_settings_key=None):
        response_delete = self.lkf_api.patch_record_list({
            "objects": [],
            "deleted_objects": ["/api/infosync/form_answer/"+str(id_record)+"/"],
            "_id":"/"
        }, jwt_settings_key=jwt_settings_key)
        print("+ response_delete:",response_delete)
        dict_response = response_delete[0][1]
        status_code = dict_response.get('status_code',0)
        return status_code

    def find_folio_autorizado(self, folio, telefono, division, tecnologia):
        record_autorizacion = self.cr_admin.find_one({
            'form_id': self.FORM_ID_AUTORIZA_CARGA_FOLIOS, 
            'deleted_at': {'$exists': False}, 
            'answers.5f93173c1a0ae8341d543f65': folio, 
            'answers.5f93173c1a0ae8341d543f66': int(telefono), 
            'answers.5f93173c1a0ae8341d543f67': division, 
            'answers.5f93173c1a0ae8341d543f68': tecnologia
        }, {'folio':1, 'answers.5fbbdac8a85f5b2ab8bb033e': 1})
        if not record_autorizacion:
            return []
        return record_autorizacion.get('answers',{}).get('5fbbdac8a85f5b2ab8bb033e', [])

    def find_tipo_tarea_catalog(self, tipo_tarea, tecnologia_orden, jwt_settings_key=None):
        mango_query = { 
            "selector": { 
                "answers": { 
                    '$and':[ 
                        { "5ec44fd1ac6a37c45828230a": { "$eq": tipo_tarea.upper() } },
                        { "5ec44ff4e065a60b1e282307": { "$eq": tecnologia_orden.upper() } } 
                    ]
                }
            }
        }
        row_catalog = self.lkf_api.search_catalog(self.CATALOG_ID_TIPOS_TAREA_FACTIBLES, mango_query, jwt_settings_key=jwt_settings_key)
        return row_catalog

    def get_all_connections(self):
        dict_to_send = {
            'url': self.settings.config['PROTOCOL'] + '://' + self.settings.config['HOST'] + '/api/infosync/connection/', 
            'method':'GET'
        }
        all_connections = self.net.dispatch(dict_to_send)
        objects = all_connections['data']
        return objects

    def get_all_contratistas_catalog(self, jwt_settings_key=None):
        records_catalog_contratistas = self.lkf_api.search_catalog(self.CATALOG_ID_CONTRATISTAS_1_0_ADMIN, self.get_mango_query_all_records(), jwt_settings_key=jwt_settings_key)
        if not records_catalog_contratistas:
            return []
        return records_catalog_contratistas

    def get_all_forms_cobranza(self):
        return {
            'fibra': {
                'metro': {'form_os': 11044,'form_lib': 23614,'form_oc': 19620},
                'sur': {'form_os': 16343,'form_lib': 44671,'form_oc': 44672},
                'norte': {'form_os': 21954,'form_lib': 50776,'form_oc': 50778},
                'occidente': {'form_os': 21953,'form_lib': 50592,'form_oc': 50594},
            },
            'cobre': {
                'metro': {'form_os': 10540,'form_lib': 29670,'form_oc': 33658},
                'sur': {'form_os': 25927,'form_lib': 45418,'form_oc': 45420},
                'norte': {'form_os': 25928,'form_lib': 50777,'form_oc': 50779},
                'occidente': {'form_os': 25929,'form_lib': 50593,'form_oc': 50595},
            }
        }

    def get_answers_form_to_form(self, form_to_migrate, answers_father):
        form_fields = self.lkf_api.get_form_id_fields(form_to_migrate)
        fields = form_fields[0]['fields']
        answers_to_record = {}
        for field in fields:
            if field.get('field_id'):
                if answers_father.get( field['field_id'] ):
                    answers_to_record[ field['field_id'] ] = answers_father[ field['field_id'] ]
        return answers_to_record

    def get_bolsa_update_communication(self, create_json, total_errors):
        dict_comunication = {
            'f1074100a010000000000a11': create_json['created']['duplicate'] or 0,
            'f1074100a010000000000a12': create_json['created']['order'] or 0,
            'f1074100a010000000000a13': total_errors,
            'f1074100a01000000000ac12': create_json['uploaded']['order'] or 0,
        }

        if create_json.get('error_file'):
            dict_comunication['f1074100a010000000000002'] = f'Existieron en la carga {total_errors} Folios con Error. Favor de revisar el archivo de folios con error'
            # dict_comunication['f1074100a010000000000003'] = create_json['error_file']
            dict_comunication['f1074100a010000000000005'] = 'error'
        else:
            dict_comunication['f1074100a010000000000002'] = 'Archivo cargado con exito'
            dict_comunication['f1074100a010000000000005'] = 'cargado'
            dict_comunication['f1074100a010000000000003'] = ""

        return dict_comunication

    def get_cambios_tecnologia(self, for_carga_pic=False):
        records_cambios = self.cr_admin.find({
            'form_id': self.FORM_ID_CAMBIO_TECNOLOGIA,
            'deleted_at': {'$exists': False}
        },{'folio': 1, 'answers': 1})
        if for_carga_pic:
            return { \
                '{}_{}'.format( rc['answers'].get('667091a65afe99d4ceba0392'), rc['answers'].get('f1054000a010000000000005') ): {'folio': rc['folio'], 'record_id': rc['_id']} \
                for rc in records_cambios \
            }
        return [ '{}_{}'.format(rc['answers'].get('667091a65afe99d4ceba0392'), rc['answers'].get('f1054000a010000000000005')) for rc in records_cambios ]

    def get_cope_in_catalogo(self, cope, find_by_abreviatura=False, jwt_settings_key=None):
        dict_to_search = { '5d641731ddd3adcc24778a9d': { '$eq': cope.upper() } }
        if find_by_abreviatura:
            dict_to_search = { '5d641731ddd3adcc24778a9e': { '$eq': cope } }
        mango_query = { "selector": { "answers": { '$and':[ dict_to_search ] } } }
        row_catalog = self.lkf_api.search_catalog(self.CATALOG_ID_COPES, mango_query, jwt_settings_key=jwt_settings_key)
        rows_copes = []
        if row_catalog:
            plain_data = ['5d641731ddd3adcc24778a9b', '5d641731ddd3adcc24778a9a', '5d641731ddd3adcc24778a9c', '5d641731ddd3adcc24778a9d', \
            '5d641731ddd3adcc24778a9e', '5d77bd96eeb32658ce522ec4']
            for info_cope in row_catalog:
                subdireccion = info_cope.get('5d641731ddd3adcc24778a9b','').lower().replace(' ','_')
                division = info_cope.get('5d641731ddd3adcc24778a9a','').lower().replace(' ','_')
                area = info_cope.get('5d641731ddd3adcc24778a9c','').lower().replace(' ','_')
                nombre_cope = info_cope.get('5d641731ddd3adcc24778a9d','').lower().replace(' ','_')
                aplica_bonificaciones = info_cope.get('60636c315093d0b3189d5779','si').lower()
                dict_info_cope_found = {
                    'subdireccion': subdireccion, 
                    'division': division, 
                    'area': area, 'cope': nombre_cope, 
                    'aplica_bonificaciones': aplica_bonificaciones, 
                    'plain_data': { i: info_cope.get(i) for i in plain_data }
                }
                rows_copes.append( dict_info_cope_found )
        return rows_copes

    def crea_directorio_temporal(self, nueva_ruta):
        # try:
        if True:
            if not os.path.exists(str(nueva_ruta)):
                os.makedirs(str(nueva_ruta),0o777)
                os.chdir(str(nueva_ruta))
            else:
                shutil.rmtree(nueva_ruta)
                os.makedirs(str(nueva_ruta),0o777)
                os.chdir(str(nueva_ruta))
            return True
        # except Exception as e:
        #     print ('############ error al crear directorio temporal=',str(e))
        #     return False

    def get_date_now(self, only_date=False):
        today_date = datetime.now(tz=timezone('America/Monterrey'))
        if only_date:
            return today_date.strftime('%Y-%m-%d')
        return today_date.strftime('%Y-%m-%d %H:%M:%S')

    def get_dias_otorgados(self, desde, festivos=False, all_days=False):
        # Modificar esta funcion para considerar las nuevas configuraciones.
        # Devolvería las fechas que apliquen para que ya no se muevan los otros scripts
        query_otorgados = {
            'form_id': 96248,
            'deleted_at': {'$exists': False},
            # 'answers.64013e0b6f74afc8b5d3318f': {'$gte': desde}, Ya no tiene caso validar esta fecha
            'answers.655e578ea69b4056a450d383': {'$nin': ['festivo']}
        }
        if festivos:
            query_otorgados['answers.655e578ea69b4056a450d383'] = 'festivo'
        if all_days:
            query_otorgados.pop('answers.655e578ea69b4056a450d383')
        
        # otorgados = cr.find(query_otorgados, {'answers': 1, 'folio': 1})
        # return [ o['answers']['64013e0b6f74afc8b5d3318f'] for o in otorgados ]
        otorgados = self.cr.aggregate([{'$match': query_otorgados},{
            '$project': {
                'tipo': '$answers.655e578ea69b4056a450d383',
                'dia_fijo': '$answers.6786eea0c35b2a22476d429f',
                'fecha': '$answers.64013e0b6f74afc8b5d3318f',
                'dia_semana': '$answers.6786eea0c35b2a22476d42a0',
                'numero_semana': '$answers.6786eea0c35b2a22476d42a1'
            }
        }])
        all_festivos = []
        current_year = self.get_date_now().split('-')[0]
        dias_semana = {
            'lunes': 0,
            'martes': 1,
            'miercoles': 2,
            'jueves': 3,
            'viernes': 4,
            'sabado': 5,
            'domingo': 6
        }
        for o in otorgados:
            fecha_parts = o['fecha'].split('-')
            fecha_nueva = "{}-{}-{}".format(current_year, fecha_parts[1], fecha_parts[2])
            fecha_festivo = o.get('fecha')
            # Si la fecha esta marcado como fijo, se toma la fecha pero con el año actual
            if o.get('dia_fijo') == 'yes':
                # Si es el primero de enero, se toma la fecha pero con el año más uno del parametro desde
                if '01-01' in o['fecha']:
                    fecha_nueva = "{}-01-01".format(int( desde.split('-')[0] ) + 1)
                fecha_festivo = fecha_nueva
            elif o.get('dia_semana') and o.get('numero_semana'):
                # Se evaluan los campos dia y numero de semana para determinar la fecha exacta para el descanso
                dia_semana_numero = dias_semana[ o.get('dia_semana') ]
                numero_semana = int(o.get('numero_semana'))
                date_nueva_fecha = self.str_to_date(fecha_nueva)
                # Ajustar al día de la semana especificado
                # Obtener el primer día del mes
                primer_dia_mes = date_nueva_fecha.replace(day=1)
                # Ajustar al día de la semana deseado
                dias_a_sumar = (dia_semana_numero - primer_dia_mes.weekday() + 7) % 7
                fecha_primera_ocurrencia = primer_dia_mes + timedelta(days=dias_a_sumar)
                # Sumar semanas según el número de semana deseado (restando 1 porque ya tenemos la primera)
                date_nueva_fecha = fecha_primera_ocurrencia + timedelta(weeks=(numero_semana-1))
                fecha_festivo = date_nueva_fecha.strftime("%Y-%m-%d")
            # Se integra la fecha festiva en la lista de festivos si es mayor o igual al parametro desde
            if self.str_to_date(fecha_festivo) >= self.str_to_date(desde):
                all_festivos.append(fecha_festivo)
        return all_festivos

    def get_element_dict(self, field):
        return {'field_id':field['field_id'], 'field_type':field['field_type'], 'label':field['label'], 'options':field['options']}

    def get_expedientes_contratistas(self, list_expedientes=None, contratistas_migrados=False, with_item_idUser=False, filter_id_contratista=None):
        match_expeds = {
            'form_id': self.FORM_ID_EXP_TECNICOS,
            'deleted_at': {'$exists': False},
            'answers.590a4761f851c20e60ac168c': 'activo',
            'answers.f1216500a010000000000001': {'$exists': True}
        }
        if list_expedientes:
            match_expeds['answers.f1216500a010000000000001'] = {'$in': list_expedientes}
        if filter_id_contratista:
            match_expeds['answers.5f344a0476c82e1bebc991d5.5f344a0476c82e1bebc991d6'] = {'$in': [str(filter_id_contratista)]}
        records_expedientes = self.cr_admin.aggregate([{
            '$match': match_expeds
        },{
            '$project': {
                'id_usuario': {'$arrayElemAt': ['$answers.5f344a0476c82e1bebc991d5.5f344a0476c82e1bebc991d6',0]},
                'expediente': '$answers.f1216500a010000000000001',
                'migrado': '$answers.6675f8383b66efd3fae66b67'
            }
        }])
        if with_item_idUser:
            return {
                int(rec_exp.get('id_usuario')): {'expediente': rec_exp['expediente'], 'migrado': rec_exp.get('migrado')} for rec_exp in records_expedientes if rec_exp.get('id_usuario')
            }
        if contratistas_migrados:
            return {
                rec_exp['expediente']: {'id_usuario': int(rec_exp.get('id_usuario')), 'migrado': rec_exp.get('migrado')} for rec_exp in records_expedientes if rec_exp.get('id_usuario')
            }
        return {
            rec_exp['expediente']: int(rec_exp.get('id_usuario')) for rec_exp in records_expedientes if rec_exp.get('id_usuario')
        }

    def get_fields_catalog(self, id_catalog):
        catalog_fields = self.lkf_api.get_catalog_id_fields(id_catalog)
        return catalog_fields.get('catalog', {}).get('fields', [])

    def get_id_os(self, division, tecnologia):
        # Diccionario que mapea las combinaciones de (division, tecnologia) a los valores deseados
        formas = {
            ('metro', 'fibra'): ( self.lkf_obj.ORDEN_SERVICIO_FIBRA, self.lkf_obj.FORMA_LIBERACION_FIBRA, self.lkf_obj.FORMA_ORDEN_COMPRA_FIBRA ),
            ('sur', 'fibra'): ( self.lkf_obj.ORDEN_SERVICIO_FIBRA_SURESTE, self.lkf_obj.FORMA_LIBERACION_FIBRA_SURESTE, self.lkf_obj.FORMA_ORDEN_COMPRA_FIBRA_SURESTE ),
            ('norte', 'fibra'): ( self.lkf_obj.ORDEN_SERVICIO_FIBRA_NORTE, self.lkf_obj.FORMA_LIBERACION_FIBRA_NORTE, self.lkf_obj.FORMA_ORDEN_COMPRA_FIBRA_NORTE ),
            ('occidente', 'fibra'): ( self.lkf_obj.ORDEN_SERVICIO_FIBRA_OCCIDENTE, self.lkf_obj.FORMA_LIBERACION_FIBRA_OCCIDENTE, self.lkf_obj.FORMA_ORDEN_COMPRA_FIBRA_OCCIDENTE ),
            ('metro', 'cobre'): ( self.lkf_obj.ORDEN_SERVICIO_COBRE, self.lkf_obj.FORMA_LIBERACION_COBRE, self.lkf_obj.FORMA_ORDEN_COMPRA_COBRE ),
            ('sur', 'cobre'): ( self.lkf_obj.ORDEN_SERVICIO_COBRE_SURESTE, self.lkf_obj.FORMA_LIBERACION_COBRE_SURESTE, self.lkf_obj.FORMA_ORDEN_COMPRA_COBRE_SURESTE ),
            ('norte', 'cobre'): ( self.lkf_obj.ORDEN_SERVICIO_COBRE_NORTE, self.lkf_obj.FORMA_LIBERACION_COBRE_NORTE, self.lkf_obj.FORMA_ORDEN_COMPRA_COBRE_NORTE ),
            ('occidente', 'cobre'): ( self.lkf_obj.ORDEN_SERVICIO_COBRE_OCCIDENTE, self.lkf_obj.FORMA_LIBERACION_COBRE_OCCIDENTE, self.lkf_obj.FORMA_ORDEN_COMPRA_COBRE_OCCIDENTE ),
        }

        # Buscar las formas correspondientes en el diccionario
        return formas.get((division, tecnologia), (None, None))

    def get_info_user_from_catalog( self, folio_o_idUser, id_catalogo, filter_by_name={} ):
        # Asignar field_id basado en el catálogo
        catalog_field_map = {
            71797: '60ca52510603f81c9db7bfdd', # Catálogo de jefes directos vinculado con plantilla
            83084: '64794444fba2ed2983d790fc', # Asignación de personal de incidencias
            59273: '5f344a0476c82e1bebc991d6' # Catalogo de Contratistas 1.0
        }
        field_id = catalog_field_map.get(id_catalogo)
        if not field_id:
            print(f"Error: ID de catálogo desconocido ({id_catalogo}).")
            return []

        mango_query = {
            "selector": {
                "answers": {
                    "$and": [{field_id: {"$eq": folio_o_idUser}}]
                }
            },
            "limit": 1,
            "skip": 0
        }
        if filter_by_name:
            mango_query["selector"]["answers"]["$and"].append( filter_by_name )
            mango_query["limit"] = 3
        print(f"*** Buscando registro en el catálogo = {id_catalogo}, query = {mango_query}")
        res = self.lkf_api.search_catalog( id_catalogo, mango_query)

        # Validar resultados
        if isinstance(res, list) and len(res) > 1:
            print("+++ +++ +++ Atención: Se encontraron múltiples registros.")

        return res if isinstance(res, list) else []

    def get_jwt_admin(self):
        return self.lkf_api.get_jwt( api_key='398bd78880b1675a4a8d06d8a89e712ad9b499fb', user='adminpclink@operacionpci.com.mx' )

    def get_mango_query_all_records(self, limit=10000):
        return { "selector": { "_id": { "$gt": None } }, "limit":limit,"skip":0 }

    def get_metadata_for_create_record(self, form_id, name_script, process='', folio_carga='', accion='Crear Registro'):
        metadata = self.lkf_api.get_metadata(form_id)
        metadata.update({ 'properties': self.lkf_obj.get_metadata_properties(name_script, accion, process=process, folio_carga=folio_carga) })
        return metadata

    # def get_metadata_properties(self, name_script, accion, process='', folio_carga=''):
    #     dict_properties = {
    #         'device_properties': {
    #             'system': 'SCRIPT',
    #             'process': process,
    #             'accion': accion,
    #             'archive': name_script
    #         }
    #     }
    #     if folio_carga:
    #         dict_properties['device_properties']['folio carga'] = folio_carga
    #     return dict_properties

    def get_nombre_tecnico(self, folio_tecnico, id_lkf=None):
        query = {'form_id':self.FORM_ID_LISTADO_TECNICOS, 'deleted_at' : {'$exists':False}, 'folio': str(folio_tecnico)}
        if id_lkf:
            query['answers.5e17ceab6a4b343af06d200e'] = id_lkf
            query.pop('folio')

        record = self.cr_admin.find_one(query, {'folio':1, 'answers':1})

        if not record:
            return ''
        
        tecnico_rec_answers = record['answers']
        nombre_tecnico_rec = tecnico_rec_answers.get('f1216500a010000000000004','')+' '+tecnico_rec_answers.get('f1216500a010000000000005','')+' '+tecnico_rec_answers.get('f1216500a010000000000006','')
        if id_lkf:
            return {"nombre": nombre_tecnico_rec, "folio_tec": record['folio']}
        return nombre_tecnico_rec
    
    def get_order_tecnology(self, header, record):
        if 'tecnologia_orden' in header:
            col = header.index('tecnologia_orden')
        if 'tecnologia_de_orden' in header:
            col = header.index('tecnologia_de_orden')
        if record[col] != '':
            tec = str(record[col]).lower().replace(u'\xa0',u' ').strip()
        else:
            tec = record[col]
        return tec

    def get_parent_id(self, user_id, all_info=False, jwt_settings_key='USER_JWT_KEY'):
        info_user = self.lkf_api.get_user_by_id(user_id, jwt_settings_key=jwt_settings_key)
        parent_info = info_user.get('parent_info',{})
        if all_info:
            return parent_info
        return parent_info.get('id',0)

    def get_permisos_contratista_from_catalog( self, id_contratista, catalog_id_contratistas=None ):
        # Se define el catalogo donde se consultan los permisos.
        # default en el catalogo del Admin
        CATALOG_CONTRATISTAS = catalog_id_contratistas or self.CATALOG_ID_CONTRATISTAS_1_0_ADMIN
        
        # El jwt quese utiliza si no mandan el catalogo a consultar se toma el jwt del admin
        jwt_to_use = 'JWT_KEY_ADMIN' if not catalog_id_contratistas else 'JWT_KEY'
        
        print('jwt_to_use consulta catalogo =',jwt_to_use)
        print('CATALOG_CONTRATISTAS =',CATALOG_CONTRATISTAS)
        
        permisos_found = {
            'carga': '', 'expediente': '', 'carga_sin_pdf': '', 'carga_sin_distometro': '',
            'libs_sin_pdf': '', 'cobre_mayor_a_300': '', 'facturar': '', 'socio_comercial': '',
            'contratista_carso': '', 'liberado_de_conecta': '', 'nombre': '', 'id_de_usuario': '',
            'plain_data': {}
        }

        field_map = {
            'carga': '5f8eff3c85a98dcae9f05dfb',
            'expediente': '5f9a41469bcab68b640bd94b',
            'carga_sin_pdf': '60be5fd1a9b326fec09658c2',
            'carga_sin_distometro': '60ec8bf3157e19fe8b393b42',
            'libs_sin_pdf': '5fac1c22380db843149b7330',
            'cobre_mayor_a_300': '613111dd9630c7d6baa69776',
            'facturar': '619e7a46c79af2f6eaf888c5',
            'socio_comercial': '614e4cd2c1770ff99f38ac33',
            'liberado_de_conecta': '63bed6a0cd55b21466e6f929',
            'nombre': '5f344a0476c82e1bebc991d7',
            'id_de_usuario': '5f344a0476c82e1bebc991d6',
            'contratista_carso': '665f70d3a7463635ed0e0b81'
        }

        plain_data_fields = [
            '5f344a0476c82e1bebc991d7', '5f344a0476c82e1bebc991db',
            '5f344a0476c82e1bebc991da', '5f344a0476c82e1bebc991d9',
            '5f344a0476c82e1bebc991d8', '673e4bc4f3e1a7191de583cf'
        ]

        mango_query = {
            "selector": {
                "answers": {
                    '$and': [{"5f344a0476c82e1bebc991d6": {"$eq": str(id_contratista)}}]
                }
            }
        }

        try:
            row_catalog = self.lkf_api.search_catalog(CATALOG_CONTRATISTAS, mango_query, jwt_settings_key=jwt_to_use)
            if not row_catalog:
                return permisos_found  # Retorna vacíos si no encontró nada

            # Tomar solo el primer registro
            r = row_catalog[0]

            # Mapear campos directos
            for key, field_id in field_map.items():
                permisos_found[key] = r.get(field_id, '')

            # Mapear plain_data
            permisos_found['plain_data'] = {fid: r.get(fid, '') for fid in plain_data_fields}

        except Exception as e:
            print(f"[ERROR] al obtener permisos del contratista {id_contratista}: {e}")
            # raise  # Lanza la excepción para que la maneje quien llame
            return False

        return permisos_found

    def get_pos_field_id_dict(self, header, form_id=None, cope='', forms_dict={}, equivalcens_map_observa={'observaciones':[]}):
        if not forms_dict.get( form_id ):
            forms_dict[ form_id ] = self.lkf_api.get_form_id_fields(form_id)
        
        form_fields = forms_dict[form_id]
        if not form_fields:
            return False
        
        header_dict = self.make_header_dict(header)
        fields = form_fields[0]['fields']

        pos_field_id = {}
        name_to_type = {
            'Clase de Servicio': 'clase',
            'Tipo de Tarea': 'tipo de tarea',
            'Num. Serie': 'num_serie',
            'usuario_reporta': 'usuario_pic_reporta'
        }
        def update_json_fields_id( name_field_to_process ):
            for folio_key in self.equivalcens_map[ name_field_to_process ]:
                folio_key = folio_key.lower().replace(' ','_')
                if folio_key in header_dict.keys():
                    pos_key = self.catch_repeted_keys(list( pos_field_id.keys() ), header_dict[folio_key])
                    pos_field_id[pos_key] = { 'scritp_type': name_to_type.get( name_field_to_process, name_field_to_process.lower().replace(' ', '_') ) }
                    if name_field_to_process == 'usuario_reporta':
                        pos_field_id[pos_key]['field_id'] = 'f1054000a030000000000111'

        for title_field in ['Folio', 'Etapa', 'Tipo', 'Clase de Servicio', 'Tipo de Tarea', 'Aerea', 'Subterranea', 'Created At', 'Contratista', 'Tecnico', \
        'Fecha de Liquidacion', 'Fecha de Asignacion', 'Cope', 'Num. Serie', 'cambio_tecnologia', 'usuario_reporta']:
            update_json_fields_id( title_field )

        if len(form_fields) > 0:
            ####
            #### For para obtener columnas esecificas del tipo de servicio
            ####
            # ---------------------------- Tienen diferencias ----------------------------
            for folio_key in equivalcens_map_observa['observaciones']:
                folio_key = folio_key.lower().replace(' ','_')
                if folio_key in header_dict.keys():
                    pos_key = self.catch_repeted_keys(list( pos_field_id.keys() ), header_dict[folio_key])
                    pos_field_id[pos_key] = {'scritp_type':'observaciones_claro_video'}
            ###
            # Se actualiza el diccionario con la informacion del campo donde se integrara la informacion
            ###
            for field in fields:
                label = field['label'].lower().replace(' ' ,'')
                label_underscore = field['label'].lower().replace(' ' ,'_')
                if label_underscore in ('clase', 'tipo', 'etapa'):
                    pos_field_id = self.update_pos_field_id(pos_field_id, position=label_underscore, field=field, scritp_type=label_underscore)
                elif label_underscore == 'tipo_de_tarea':
                    pos_field_id = self.update_pos_field_id(pos_field_id, position='tipo de tarea', field=field, scritp_type='tipo de tarea')
                elif label_underscore == 'activacion_claro_video':
                    pos_field_id = self.update_pos_field_id(pos_field_id, position=header_dict.get(label_underscore, -1), field=field, scritp_type='activacion_claro_video')
                if label in header_dict.keys():
                    pos_field_id = self.update_pos_field_id(pos_field_id, position=header_dict.get(label, -1), field=field)
                elif label_underscore in header_dict.keys():
                    pos_field_id = self.update_pos_field_id(pos_field_id, position=header_dict.get(label_underscore, -1), field=field)
                elif field['label'] in self.equivalcens_map.keys():
                    for eqv_label in self.equivalcens_map[field['label']]:
                        header_lable = eqv_label.lower().replace(' ' ,'')
                        header_lable_under = eqv_label.lower().replace(' ' ,'_')
                        if header_lable in header_dict.keys():
                            pos_field_id = self.update_pos_field_id(pos_field_id, position=header_dict.get(header_lable, -1), field=field)
                        elif header_lable_under in header_dict.keys():
                            pos_field_id = self.update_pos_field_id(pos_field_id, position=header_dict.get(header_lable_under, -1), field=field)
        return pos_field_id

    def get_query_plantilla(self, extra_filters={}):
        query_plantilla = {
            'form_id': self.FORM_ID_PLANTILLA_EMPRESARIAL,
            'deleted_at': {'$exists': False},
            'answers.5d5ec9a7c82d82ff8dbb41ac': 'activo'
        }
        if extra_filters:
            query_plantilla.update(extra_filters)
        return query_plantilla
    
    def get_record_folio(self, header_dict):
        folio_col = header_dict.get('folio', header_dict.get('folio_pisa') )
        if folio_col is None:
            for folio_key in self.equivalcens_map['Folio']:
                folio_key = folio_key.lower().replace(' ','_')
                if folio_key in header_dict:
                    folio_col = header_dict[folio_key]
                    break
        return folio_col
        # if folio_col is None:
        #     return None
        # return str(record[folio_col])

    def get_record_from_db(self, forma_id, folio):
        rec = self.cr.find_one({
            'form_id': forma_id, 'folio': folio, 'deleted_at': {'$exists': False}
        },{'form_id': 1, 'folio': 1, 'other_versions': 1, 'answers': 1})
        return rec

    def get_records_query(self, form_id, folio):
        query = {
            'form_id': form_id, 
            'deleted_at': {'$exists':False},
            'folio': {'$in': folio} if (type(folio) == list) else folio
        }
        select_columns = {'folio':1,'user_id':1,'form_id':1, 'answers':1,'_id':1,'connection_id':1}
        return query, select_columns

    def get_records_existentes(self, form_id, folios, extra_params={}, os_with_phone=False):
        query, select_columns = self.get_records_query(form_id, folios)
        query.update(extra_params)
        record = self.cr.find(query, select_columns)
        if os_with_phone:
            existentes = { '{}{}'.format(rec['folio'], rec['answers'].get('f1054000a010000000000005')): rec for rec in record }
            return existentes
        existentes = {rec['folio']:rec for rec in record}
        return existentes

    def get_today(self, str_format='%Y-%m-%d'):
        dtoday = datetime.today()
        new_tz = timezone('America/Monterrey')
        today_timezone = new_tz.fromutc(dtoday)
        return today_timezone.strftime(str_format)

    def header_without_acentos(self, header):
        new_header = []
        for h in header:
            new_header.append( self.quit_acentos( h ) )
        return new_header

    def make_cr_connection( self, id_connection ):
        from pci_get_connection_db import CollectionConnection
        colection_connection = CollectionConnection(id_connection, self.settings)
        return colection_connection.get_collections_connection()

    def make_excel_file(self, rows, content_sheets={}):
        #rows = make_array(orders)
        date = datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")
        file_name = "/tmp/output_%s.xlsx"%(date)
        if content_sheets:
            my_excel = pyexcel.get_book(bookdict=content_sheets)
            my_excel.save_as(file_name)
        else:
            pyexcel.save_as(array=rows, dest_file_name=file_name)
        return file_name

    def make_header_dict(self, header):
        ### Return the directory with
        ### the column name : column number
        header_dict = {}
        for position in range(len(header)):
            header_dict[ str( header[ position ] ).lower().replace(' ' ,'_') ] = position
        return header_dict

    def match_answers_record_with_catalog( self, id_catalog, answers_record, fields_catalog_to_process=None ):
        """
        Empareja respuestas de un registro con las del catálogo correspondiente.

        Args:
            id_catalog (int): ID del catálogo.
            answers_record (dict): Respuestas del registro.

        Returns:
            dict: Respuestas formateadas para el catálogo.
        """
        fields = fields_catalog_to_process if fields_catalog_to_process else self.get_fields_catalog(id_catalog)
        answers_to_catalog = {}
        for f in fields:
            id_field = f.get('field_id')
            options = { o.get('value'): o.get('label') for o in f.get('options',[]) }
            
            # Se valida que el answers tenga informacion en el campo que se esta procesando
            if answers_record.get(id_field):
                
                # Si hay opciones entonces se entiende que es un campo de tipo Respuesta multiple
                # Por tanto se debe optener el valor que le corresponde a la opción
                if options:
                    answers_to_catalog[id_field] = options[ answers_record[id_field] ]
                    continue
                
                # Cualquier otro campo que no sea de opciones se pasa tal cual al catalog
                answers_to_catalog[id_field] = answers_record[id_field]
        return answers_to_catalog

    def merge_json(self, dict_a, dict_b, header, form_id):
        record_errors = []
        error_file = False
        AA = dict_a['uploaded']
        BB = dict_b['uploaded']
        uploaded = {x: AA.get(x, 0) + BB.get(x, 0) for x in set(AA).union(BB)}
        AA = dict_a['created']
        BB = dict_b['created']
        created = {x: AA.get(x, 0) + BB.get(x, 0) for x in set(AA).union(BB)}
        if dict_a['error_file_records']:
            record_errors +=  dict_a['error_file_records']
        if dict_b['error_file_records']:
            if dict_a['error_file_records']:
                for error in dict_b['error_file_records']:
                    duplicate = False
                    for rec in dict_a['error_file_records']:
                        if rec[:15] == error[:15]:
                            duplicate = True
                    if not duplicate:
                        record_errors += [error,]
            else:
                record_errors +=  dict_b['error_file_records']
        if record_errors:
            error_file = self.upload_error_file(header + ['error',],  record_errors, form_id)
        res = {'uploaded':uploaded,'error_file':error_file, 'created':created }
        return res

    def query_folio_preorder(self, form_id , folio, telefono, area):
        query = {'form_id':  form_id, 'deleted_at' : {'$exists':False}, 
            'folio': str(folio),
            'answers.f1054000a010000000000005': telefono,
            'answers.f1054000a0100000000000a2': area
        }
        select_columns = {'folio':1, 'answers':1, 'connection_id':1, 'user_id':1}
        return query, select_columns

    def quit_acentos(self, cadena):
        dict_replace = { 'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u' }
        for a, aa in dict_replace.items():
            cadena = cadena.replace( a, aa )
        return cadena

    def get_num_rows_in_xls(self, xls_url):
        """
        Lee la cantidad de renglones que hay un excel

        Args:
            xls_url (str): URL del excel en backblaze

        Return:
            num_rows (int): Numero de lineas encontradas en el archivo
        """
        # Descargar el archivo
        response = requests.get(xls_url)
        response.raise_for_status()

        # Cargar el archivo excel desde la memoria
        excel_file = BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_file)

        # Accedemos a la hoja
        hoja = wb.active

        # Numero de lineas
        num_rows = hoja.max_row
        return num_rows

    def read_file(self, file_url):
        """
        Lee un documento excel y regresa el header y renglones

        Args:
            file_url (str): URL del excel

        Return:
            header (list): Primer renglon del excel que se toma como cabecera
            all_records (list): Renglones del excel
        """
        max_row_accepted = 5000
        if self.get_num_rows_in_xls(file_url) > max_row_accepted:
            return {'error': f'El excel rebasa el límite de renglones permitidos {max_row_accepted}. Favor de revisar'}, None
        
        sheet = pyexcel.get_sheet(url = file_url)
        all_records = sheet.array
        header = [h.lower().strip().replace(' ', '_') for h in all_records.pop(0)]
        return header, all_records

    def rename_file_extract( self, actual_name ):
        # actual_name = actual_name.decode('latin-1')
        # actual_name = actual_name.encode('ascii', 'replace')
        # actual_name = actual_name.replace('?', '')

        # Suponiendo que actual_name es de tipo bytes inicialmente
        if isinstance(actual_name, bytes):
            actual_name = actual_name.decode('latin-1')  # bytes -> str

        # Convertir a ASCII, reemplazar caracteres no ASCII por '?'
        actual_name = actual_name.encode('ascii', 'replace').decode('ascii')

        # Eliminar los caracteres '?' resultantes
        actual_name = actual_name.replace('?', '')

        return actual_name

    def review_psr_exists(self, folio, telefono):
        record_psr = self.cr_admin.find_one({
            'form_id': self.FORM_ID_PSR, 'deleted_at': {'$exists': False},
            'answers.667091a65afe99d4ceba0392': folio,
            'answers.f1054000a010000000000005': telefono
        }, {'folio': 1, 'answers': 1})
        return record_psr

    def search_for_connection(self, connection_email, form_id, dir_form_connection):
        print('connection_email = ',connection_email)
        print('form_id = ',form_id)

        if not dir_form_connection.get(form_id):
            dir_form_connection[form_id] = self.lkf_api.get_form_connections(form_id)
        print('dir_form_connection = ',dir_form_connection)
        form_shareds = dir_form_connection[form_id]
        shared_connections = form_shareds.get('connections',{})
        for connection in shared_connections:
            if connection_email == connection['email']:
                connection['connection_id'] = connection['id']
                return connection

        if form_shareds.get('coworkers'):
            for connection_id, coworkers in form_shareds.get('coworkers',{}).items():
                if coworkers:
                    for worker in coworkers:
                        if connection_email == worker['email']:
                            worker['connection_id'] = connection_id
                            return worker
        return {}

    def set_metros_bajante(self, mts_orden, tecnologia_orden, hibrida, field_id_metraje, error = []):
        res = {}
        if mts_orden > 300 and tecnologia_orden == 'fibra':
            if mts_orden > 1125:
                error.append('Error de metraje solo es valido hasta 1125')
                return res

            res.update({field_id_metraje:'300'})

            res.update({'f1054000a0200000000000d7':str(int(300))})
            res.update({'f1054000a020000000000bd7':str(int(mts_orden - 300))})
        else:
            if tecnologia_orden == 'cobre':
                res.update({
                    field_id_metraje: 350 if int( mts_orden ) > 350 else int( mts_orden )
                })
            else:
                if hibrida:
                    mts_orden = self.adjust_metros_fibra(mts_orden)
                res.update({field_id_metraje:str(int(mts_orden))})

        return res

    def set_status_values(self, answer, is_update):
        custom_answer = {}
        estatus_orden = answer.get('f1054000a030000000000002',False)
        if estatus_orden:
            estatus_orden = estatus_orden.lower().replace(' ','_')
        if not estatus_orden or estatus_orden not in ['pendiente', 'cita_programada', 'liquidada', 'cliente_no_se_localiza', 'objetada', 'pendiente_de_trabajo', 'reasignada']:
            custom_answer['f1054000a030000000000002'] = 'liquidada'
        if not is_update:
            custom_answer['f1054000a030000000000012'] = 'pendiente'
        return custom_answer

    def send_notification_email(self, name_process, msg, name_script='', folio_carga='', jwt_settings_key=None):
        metadata_email_error = self.lkf_api.get_metadata( form_id=self.FORM_ID_EMAIL_ERRORES )
        metadata_email_error['properties'] = self.lkf_obj.get_metadata_properties(name_script, '', process=name_process, folio_carga=folio_carga)
        metadata_email_error['answers'] = {
            '668ff5ef62f58d0b8dca8ba7': name_process,
            '668ff5ef62f58d0b8dca8ba8': msg
        }
        res_create_msg_error = self.lkf_api.post_forms_answers(metadata_email_error, jwt_settings_key=jwt_settings_key)

    # def set_status_proceso(self, current_record, status_set, field_status='6291068e438a7381b949954c', jwt_settings_key=None):
    #     current_record['answers'][field_status] = status_set
    #     self.lkf_api.patch_record(current_record, jwt_settings_key=jwt_settings_key)

    def set_status_proceso(self, current_record, record_id, status_set, msg='', field_status='5f10d2efbcfe0371cb2fbd39', field_msg='5fd05319cd189468810100c9' ):
        current_record['answers'][field_status] = status_set
        current_record['answers'][field_msg] = msg
        return self.lkf_api.patch_record(current_record, record_id,  jwt_settings_key='USER_JWT_KEY')

    def str_to_date(self, val_str, format_to_date='%Y-%m-%d'):
        return datetime.strptime(val_str, format_to_date)

    def str_to_float(self, val_str):
        return float( val_str.strip().replace(',','').replace('$','') )

    # def list_to_str(self, list_to_proccess, separator=', ', show_empty=False):
    #     str_return = ''
    #     if show_empty:
    #         str_return += separator.join([a for a in list_to_proccess])
    #     else:
    #         str_return += separator.join([a for a in list_to_proccess if a])
    #     return str_return

    def unlist(self, arg):
        if type(arg) == list and len(arg) > 0:
            return self.unlist(arg[0])
        return arg

    def update_create_json(self, create_json, this_record):
        if this_record.get('create') and this_record['create']:
            create_json['created']['total'] += 1
            create_json['created']['order'] += 1
        if this_record.get('update') and this_record['update']:
            create_json['uploaded']['total'] += 1
            create_json['uploaded']['order'] += 1
        return create_json

    def update_pos_field_id(self, pos_field_id, position, field, scritp_type=''):
        if pos_field_id.get(position):
            pos_field_id[position].update(self.get_element_dict(field))
        else:
            pos_field_id[position] = self.get_element_dict(field)

        if scritp_type:
            pos_field_id[position]['scritp_type'] = scritp_type

        if not pos_field_id[position].get('scritp_type'):
            pos_field_id[position]['scritp_type'] = ''
        return pos_field_id
    
    def upload_error_file(self, header, record_errors, form_id, file_field_id='f1074100a010000000000003', content_sheets={}):
        record_errors.insert(0,header)
        archivo_file_name = self.make_excel_file(record_errors, content_sheets=content_sheets)
        csv_file = open(archivo_file_name,'rb')
        csv_file_dir = {'File': csv_file}
        # try:
        if True:
            upload_data = {'form_id': form_id, 'field_id': file_field_id}
            print('upload_data===',upload_data)
            upload_url = self.lkf_api.post_upload_file(data=upload_data, up_file=csv_file_dir,  jwt_settings_key='USER_JWT_KEY')
            print("******************** upload_url ERROR:",upload_url)
        # except:
        #     return "No se pudo generar el archivo de error "

        csv_file.close()
        try:
            file_url = upload_url['data']['file']
            # file_date = time.strftime("%Y_%m_%d")
            file_date = datetime.now().strftime("%Y_%m_%d")
            error_file = {file_field_id: {'file_name': f'Errores de Carga {file_date}.xlsx',
                                                'file_url':file_url}}
        except KeyError:
            print('could not save file Errores')
            return 'No se pudo guardar el archivo, favor de reprocesar'
        return error_file

    def upload_file_xls_formated( self, header, records_xls, form_id, file_field_id, columns_format={}, new_name='', width_column_full=False, header_color=False, incidencias_col_yellow=[], \
    paint_rows={}, cells_colors={}, jwt_settings_key='USER_JWT_KEY' ):
        date = datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")
        default_name_xls = f'output_{date}.xlsx'
        file_name = f"/tmp/{default_name_xls}"
        wrkb = openpyxl.Workbook()
        sheet = wrkb.active
        # Number of sheets in the workbook (1 sheet in our case)
        ws = wrkb.worksheets[0]
        records_xls.insert(0, header)
        
        pos_row = 1
        for row_xls in records_xls:
            ws.append(row_xls)
            if pos_row > 1:
                for letter in columns_format:
                    ws[ '{}{}'.format(letter, pos_row) ].number_format = columns_format[letter]
            if pos_row > 2:
                for col in incidencias_col_yellow:
                    letter_colum = get_column_letter(col)
                    ws[ '{}{}'.format(letter_colum, pos_row) ].fill = PatternFill("solid", start_color="00FFCC00") # Font(color="00FFCC00")
            if paint_rows.get(pos_row):
                config_paint_row = paint_rows[pos_row]
                if config_paint_row['paint_type'] == 'cell':
                    letter_colum = get_column_letter(config_paint_row['column'])
                    ws[ '{}{}'.format(letter_colum, pos_row) ].fill = PatternFill("solid", start_color=config_paint_row['color'])
            if cells_colors and cells_colors.get(pos_row):
                for color4cell in cells_colors[pos_row]:
                    cols4color = cells_colors[pos_row][color4cell]
                    for cellColor in cols4color:
                        letter_colum = get_column_letter(cellColor)
                        ws[ '{}{}'.format(letter_colum, pos_row) ].fill = PatternFill("solid", start_color=color4cell)
            pos_row += 1
        # Ancho de la columna siempre el máximo para que se vea la información completa
        if width_column_full:
            for column_cells in ws.columns:
                length = max(len(self.as_text(cell.value)) for cell in column_cells)
                num_column = column_cells[0].column
                #letter_colum = get_column_letter(num_column)
                ws.column_dimensions[num_column].width = length
                #ws.column_dimensions[ num_column ].auto_size = True
        if header_color:
            header_style = NamedStyle(name="header_style")
            header_style.font = Font(bold=True, name='Calibri')
            cell_color_header = 'B0E0E6'
            if type(header_color) == dict:
                if header_color.get('color_font'):
                    header_style.font = Font(bold=True, name='Calibri', color=header_color['color_font'])
                if header_color.get('color'):
                    cell_color_header = header_color['color']
            wrkb.add_named_style(header_style)
            header_style.fill = PatternFill("solid", start_color=cell_color_header)
            for column_cells in ws.columns:
                num_column = column_cells[0].column
                ws[ '{}1'.format( num_column ) ].style = header_style
        #ws['A3'].number_format = '#,##0.00'
        #ws['B3'].number_format = '0.00%'
        wrkb.save(file_name)
        xls_file = open(file_name, 'rb')
        xls_file_dir = {'File': xls_file}
        upload_url = self.lkf_api.post_upload_file(data={'form_id': form_id, 'field_id': file_field_id}, up_file=xls_file_dir, jwt_settings_key=jwt_settings_key)
        print('=========== upload_url:: ',upload_url)
        xls_file.close()
        try:
            file_url = upload_url['data']['file']
            error_file = {file_field_id: {'file_name': default_name_xls, 'file_url':file_url}}
        except KeyError:
            print('could not save file Errores')
            return 'No se pudo guardar el archivo, favor de reprocesar'
        if new_name:
            error_file[ file_field_id ].update({ 'file_name': new_name })
        return error_file

    def uploaded_by_other_connection(self, record, connection_id):
        #Verifica q no haya sido cargada por otra conexion
        if record.get('connection_id'):
            return 'update' if connection_id == record['connection_id'] else 'by_other'
        return 'assigne'

    def validate_record_status(self, record):
        if record['answers'].get('f1054000a030000000000012'):
            if record['answers'].get('f1054000a030000000000012', '') in ('pendiente', 'reintento'):
                return True
            if record['answers'].get('f1054000a030000000000012', '') == 'estimacion' and record['answers'].get('5fc9269ce5363f7e3b9e3867', 'no') != 'no':
                return True
            return False
        if record['answers'].get('f1054000a030000000000002', '') == 'liquidada':
            return True
        return False