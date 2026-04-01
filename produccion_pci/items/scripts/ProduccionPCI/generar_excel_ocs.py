# -*- coding: utf-8 -*-
"""
Script que se ejecuta al terminar el proceso de Liberaciones de Orden de Compra
para generar el archivo Excel de OCS y enviarlo por correo a los contratistas
"""
import sys, simplejson
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment

from collections import defaultdict

from produccion_pci_utils import Produccion_PCI
from account_settings import *

class GenerarExcelOcs( Produccion_PCI ):
    """docstring for GenerarExcelOcs"""
    def __init__(self, settings, sys_argv=None, use_api=False):
        super().__init__(settings, sys_argv=sys_argv, use_api=use_api)
        self.forms_ocs_ftth = [142287, 148227]
        self.form_id_email = 149857
        self.field_xls_email = '69c98863121f86532adeb490'

        self.filters = self.data.get('data', {})

        self.fecha_periodo = self.get_fecha_periodo()

    def get_fecha_periodo(self):
        # Se obtiene el dato de cuando se ejecuta desde el script de Orden de Compra
        if self.filters.get('fecha_corte'):
            return self.filters.get('fecha_corte')

        # Se obtiene de data cuando se ejecuta directo en la pantalla de Scripts
        if self.data.get('fecha_corte'):
            return self.data.get('fecha_corte')

        return None
    
    def get_query_ocs(self):
        query = [
            {
                '$match': {
                    'form_id': { '$in': [
                        self.FORMA_ORDEN_COMPRA_FIBRA_OCCIDENTE, 
                        self.FORMA_ORDEN_COMPRA_FIBRA_TELNOR, 
                        self.FORMA_ORDEN_COMPRA_COBRE_OCCIDENTE,
                        self.FORMA_ORDEN_COMPRA_COBRE_TELNOR,
                    ] },
                    'deleted_at': { '$exists': False },
                    'created_at': { 
                        '$gte': datetime.strptime(f"{self.fecha_periodo} 00:00:00", '%Y-%m-%d %H:%M:%S') ,
                        '$lt': datetime.strptime(f"{self.fecha_periodo} 23:59:59", '%Y-%m-%d %H:%M:%S') 
                    },

                    # 'connection_id': 29125 # Test!
                }
            },

            # Extraer campos necesarios desde raíz
            {
                '$project': {
                    'connection_id': 1,
                    'connection_name': 1,
                    'connection_email': 1,
                    'form_id': 1,
                    'folio_oc': '$folio',
                    'record_id_oc': '$_id',
                    'retencion_resico': '$answers.621fe75ad98a2471ed6308f8',
                    'iva': '$answers.f19620000000000000000f7b',
                    'total': '$answers.f19620000000000000000fc7',
                    'descuento_prestamo': '$answers.6938be2954a3f5b19d30fa0a',
                    'descuento_adeudo': '$answers.6938be2954a3f5b19d30fa0b',
                    'descuento_vehiculo': '$answers.6938be2954a3f5b19d30fa0c',
                    'descuento_ahorro': '$answers.6938be2954a3f5b19d30fa0d',
                    'detail_source': {
                        '$ifNull': ["$answers.f1962000000000000000fc10", []]
                    }
                }
            },

            # Mapear detail_folios correctamente
            {
                '$project': {
                    'connection_id': 1,
                    'connection_name': 1,
                    'connection_email': 1,
                    'form_id': 1,
                    'detail_folios': {
                        '$map': {
                        'input': '$detail_source',
                        'as': 'df',
                        'in': {
                            'folio': '$$df.f19620000000000000001fc1',
                            'telefono': '$$df.f19620000000000000001fc2',
                            'tipo_tarea': '$$df.68f6a337764a6c7697770f8c',
                            'trabajo_realizado': '$$df.f19620000000000000001fc4',
                            'reparacion_instalaciones': '$$df.672cfb27388bff96a3650582',
                            'incentivo_psr': '$$df.672cfb27388bff96a3650581',
                            'bono_productividad': '$$df.68f6a337764a6c7697770f8d',
                            'importe_a_pagar': "$$df.f1962000000000000001fc10",
                            'folio_oc': '$folio_oc',
                            'record_id_oc': '$record_id_oc',
                            'retencion_resico': '$retencion_resico',
                            'iva': '$iva',
                            'total': '$total',
                            'descuento_prestamo': '$descuento_prestamo',
                            'descuento_adeudo': '$descuento_adeudo',
                            'descuento_vehiculo': '$descuento_vehiculo',
                            'descuento_ahorro': '$descuento_ahorro',
                        }
                        }
                    }
                }
            },

            # Agrupar por connection_id + form_id
            {
                '$group': {
                '_id': {
                    'connection_id': '$connection_id',
                    'form_id': '$form_id'
                },
                'connection_name': { '$first': '$connection_name' },
                'connection_email': { '$first': '$connection_email' },
                'detail_folios': { '$push': '$detail_folios' }
                }
            },

            # Aplanar arrays
            {
                '$project': {
                'connection_id': '$_id.connection_id',
                'form_id': '$_id.form_id',
                'connection_name': 1,
                'connection_email': 1,
                'detail_folios': {
                    '$reduce': {
                        'input': '$detail_folios',
                        'initialValue': [],
                        'in': { '$concatArrays': ['$$value', '$$this'] }
                    }
                }
                }
            },

            # Agrupar por connection_id
            {
                '$group': {
                '_id': '$connection_id',
                'connection_name': { '$first': '$connection_name' },
                'connection_email': { '$first': '$connection_email' },
                'forms': {
                    '$push': {
                        'form_id': '$form_id',
                        'detail_folios': '$detail_folios'
                    }
                }
                }
            },

            # Formato final
            {
                '$project': {
                    '_id': 0,
                    'connection_id': '$_id',
                    'connection_name': 1,
                    'connection_email': 1,
                    'forms': 1
                }
            }
        ]
        return query
    
    def get_type_folio( self, folio_detail ):
        if folio_detail.get('trabajo_realizado') == 'PSR':
            if folio_detail.get('incentivo_psr'):
                return 'incentivo_psr'
            return 'quejas_psr'
        
        tipo_tarea = folio_detail.get('tipo_tarea') or ''
        tipo_os = tipo_tarea[:2]
        if tipo_os == 'A4':
            return 'a4'
        
        if tipo_os in ('TS', 'TI'):
            return 'migraciones'
        
        return None
    
    def process_importes( self, list_importes ):
        if not list_importes:
            return 0, 0, 0, set()
        
        total_importe, total_bono, count_bono = 0, 0, 0
        ordenes_compra = set()
        for item_importe in list_importes:
            bono = item_importe.get('bono') or 0
            if bono:
                total_bono += bono
                count_bono += 1
            
            total_importe += ( item_importe.get('importe', 0) - bono )
            folio_oc = item_importe.get('folio_oc')
            record_id_oc = item_importe.get('record_id_oc')
            ordenes_compra.add( f"{folio_oc} | {record_id_oc}" )
        
        return total_importe, total_bono, count_bono, ordenes_compra
    
    def add_coma(self, snum):
        return "{:,.2f}".format( float(snum) )
    
    def str_to_float(self, val):
        if not val:
            return 0
        
        if isinstance(val, float):
            return val
        
        return float( val.strip().replace(',', '').replace('$', '') )

    def process_link(self, val):
        val = val.replace('LINK: ','')
        fol_oc, id_oc = val.split(' | ', 1)
        return fol_oc, f"https://app.linkaform.com/#/records/detail/{id_oc}"
    
    def generar_xls(self, records_xls):
        date = datetime.now().strftime("%Y_%m_%d_%H_%M_%S_%f")
        default_name_xls = f'output_{date}.xlsx'
        file_name = f"/tmp/{default_name_xls}"
        # print('file_name =',file_name)
        wrkb = openpyxl.Workbook()


        # Se preparan los estilos para las celdas
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, name='Calibri')
        wrkb.add_named_style(header_style)
        header_style.fill = PatternFill("solid", start_color='DBDBDB')
        
        relleno_purpura = PatternFill(start_color='F2D5F2', end_color='F2D5F2', fill_type='solid')

        relleno_crema = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        fuente_negrita = Font(bold=True, name='Calibri')
        centrado = Alignment(horizontal='center', vertical='center')


        sheet = wrkb.active
        # Number of sheets in the workbook (1 sheet in our case)
        ws = wrkb.worksheets[0]
        
        pos_row = 1
        for row_xls in records_xls:
            ws.append(row_xls)

            # Se configuran los posibles Links que vengan en las celdas
            for pos_celda, val_celda in enumerate(row_xls):
                if isinstance(val_celda, str) and 'LINK:' in val_celda:
                    str_link, url_link = self.process_link(val_celda)
                    num_fila = ws.max_row
                    num_col = pos_celda + 1
                    celda_link = ws.cell(row=num_fila, column=num_col)
                    celda_link.value = str_link
                    celda_link.hyperlink = url_link
                    celda_link.font = Font(color="0000FF", underline="single", name='Calibri')
        
        # 2. Alinear a la derecha toda la columna C
        # Recorremos desde la fila 1 hasta la última fila con datos
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 19
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 15

        # Estino del titulo donde esta el nombre del contratista
        titulo_nombre_contratista = ws['A1']
        titulo_nombre_contratista.style = header_style
        ws.merge_cells('A1:C1')
        titulo_nombre_contratista.alignment = Alignment(horizontal='center', vertical='center')

        # Area titulo TOTAL DESCUENTO
        titulo_total_descuento = ws['A13']
        titulo_total_descuento.value = 'TOTAL DESCUENTO'
        ws.merge_cells('A13:A20')
        titulo_total_descuento.alignment = Alignment(horizontal='center', vertical='center')
        titulo_total_descuento.font = fuente_negrita
        titulo_total_descuento.fill = relleno_purpura

        # Celdas donde lleva negrita
        for l in ['A', 'B', 'C', 'D', 'E']:
            for c in [2, 10, 20, 22, 24, 25, 26, 27, 29]:
                ws[ f"{l}{c}" ].font = fuente_negrita
            
            # Renglon titulo del area donde se ponen los renglones de NOMINA
            ws[ f"{l}29" ].fill = relleno_crema
            ws[ f"{l}29" ].alignment = centrado
        
        wrkb.save(file_name)

        # Se carga el excel generado a Linkaform
        xls_file = open(file_name, 'rb')
        xls_file_dir = {'File': xls_file}
        resp_xls_upload = self.lkf_api.post_upload_file(data={'form_id': self.form_id_email, 'field_id': self.field_xls_email}, up_file=xls_file_dir)
        print('- resp_xls_upload:',resp_xls_upload)
        xls_file.close()
        try:
            file_url = resp_xls_upload['data']['file']
            return {self.field_xls_email: [{'file_name': 'Ordenes de Compra.xlsx', 'file_url':file_url}]}
        except KeyError:
            print('Error al cargar el excel')
            return None

    def get_str_descuentos(self, detail):
        prestamo = detail.get('descuento_prestamo', 0)
        adeudo = detail.get('descuento_adeudo', 0)
        vehiculo = detail.get('descuento_vehiculo', 0)
        ahorro = detail.get('descuento_ahorro', 0)
        
        return f"{detail.get('folio_oc')} | prestamo={prestamo} adeudo={adeudo} vehiculo={vehiculo} ahorro={ahorro}"

    def get_str_totales(self, detail):
        retencion_resico = detail.get('retencion_resico', 0)
        iva = detail.get('iva', 0)
        total = detail.get('total', 0)

        return f"{detail.get('folio_oc')} | retencion_resico={retencion_resico} iva={iva} total={total}"

    def get_records_nomina(self, connection_id, folios_ocs):
        records_nomina = self.cr.aggregate([
            {
                '$match': {
                    'form_id': 142281,
                    'deleted_at': {'$exists': False},
                    'answers.68ef3a6f7b3f032ba9879047': 'aplicado',
                    'answers.68fbbf8eb7c9c274a2a24955.5f344a0476c82e1bebc991d6': {'$in': [connection_id, str(connection_id)]},
                    'answers.68f087bb782a7cd1f064d8f1': {'$in': folios_ocs}
                }
            },
            {'$unwind': '$answers.696fb1adec32245b5a4dc4b6'},
            {
                '$project': {
                    'tipo_pago': '$answers.696fb1adec32245b5a4dc4b6.696fb27591fad284644dc4af',
                    'neto_pagado': {
                        '$toDouble': '$answers.696fb1adec32245b5a4dc4b6.696fb27591fad284644dc4ad'
                    },
                    'item': {
                        'nombre': '$answers.696fb1adec32245b5a4dc4b6.696fb2db8d994178f37cdb6b.696fb2db8d994178f37cdb6c',
                        'neto_pagado': '$answers.696fb1adec32245b5a4dc4b6.696fb27591fad284644dc4ad',
                        'observaciones': '$answers.696fb1adec32245b5a4dc4b6.696fb27591fad284644dc4ae',
                        'fecha_inicio': '$answers.696fb1adec32245b5a4dc4b6.696fb27591fad284644dc4b0',
                        'fecha_fin': '$answers.696fb1adec32245b5a4dc4b6.696fb27591fad284644dc4b1',
                    }
                }
            },
            {
                '$group': {
                    '_id': '$tipo_pago',
                    'total_neto_pagado': {'$sum': '$neto_pagado'},
                    'items': {'$push': '$item'}
                }
            },
            {
                '$project': {
                    '_id': 0,
                    'tipo_pago': '$_id',
                    'total_neto_pagado': 1,
                    'items': 1
                }
            }
        ])

        return {
            nomina['tipo_pago']: nomina
            for nomina in records_nomina
        }

    def parse_descuentos(self, descuentos_list):
        """
        Convierte lista de strings en:
        - dict acumulado de descuentos
        - lista de folios
        """
        map_descuentos = defaultdict(float)
        folios_ocs = []

        for raw in descuentos_list:
            try:
                folio, descuentos_str = raw.split(' | ', 1)
                folios_ocs.append(folio)

                for item in descuentos_str.split():
                    name, value = item.split('=')
                    map_descuentos[name] += self.str_to_float(value)

            except ValueError:
                # Manejo de errores por formato inesperado
                print(f"Formato inválido: {raw}")
                continue

        return map_descuentos, folios_ocs

    def process_descuentos(self, data_descuentos, kwargs_fields):
        if not data_descuentos:
            return []
        # for name_contratista, data_descuentos in group_Descuentos.items():
        #     print(f'\n === === === {name_contratista} === === ===')
        
        GRUPOS_NOMINA = ['nómina', 'imss', 'infonavit']
        GRUPOS_EXTRA = ['prestamo', 'adeudo', 'vehiculo', 'ahorro']

        rows_descuentos, rows_tecnicos = [], [['TECNICO', 'TOTAL', 'OBSERVACIONES', 'PERIODO', 'CONCEPTO']]

        descuentos = data_descuentos.get('descuentos')
        map_descuentos, folios_ocs = self.parse_descuentos(descuentos)

        # print(f"     OCs = {folios_ocs} conexion= {data_descuentos.get('id')}")
        print('     ',dict(map_descuentos))
        nominas = self.get_records_nomina( data_descuentos.get('id'), folios_ocs )
        # print( '        ',nominas )
        total_descuentos = 0
        for grupo_descuento in GRUPOS_NOMINA:
            data_nominas = nominas.get(grupo_descuento, {})
            descuento = data_nominas.get('total_neto_pagado', 0)
            total_descuentos += descuento
            rows_descuentos.append(['', grupo_descuento.upper(), self.add_coma( descuento )])

            kwargs_fields.setdefault('69cacdc9dcfb7b9636860e1d', []).append({
                '69cace21a461e2086b34e260': grupo_descuento.upper(),
                '69cace21a461e2086b34e261': round(descuento, 2)
            })

            for item_tecnico in data_nominas.get('items', []):
                rows_tecnicos.append([
                    item_tecnico.get('nombre'),
                    item_tecnico.get('neto_pagado', 0),
                    item_tecnico.get('observaciones', ''),
                    f"DEL {item_tecnico.get('fecha_inicio', '')} AL {item_tecnico.get('fecha_fin', '')}",
                    grupo_descuento.upper()
                ])

                kwargs_fields.setdefault('69cace87bba91641eb8c3f6f', []).append({
                    '69cacee4056f23facbe1dd19': item_tecnico.get('nombre'),
                    '69cacee4056f23facbe1dd1a': item_tecnico.get('neto_pagado', 0),
                    '69cacee4056f23facbe1dd1b': item_tecnico.get('observaciones', ''),
                    '69cacee4056f23facbe1dd1c': f"DEL {item_tecnico.get('fecha_inicio', '')} AL {item_tecnico.get('fecha_fin', '')}",
                    '69cacee4056f23facbe1dd1d': grupo_descuento.upper()
                })

        for grupo_descuento in GRUPOS_EXTRA:
            descuento = map_descuentos.get(grupo_descuento, 0)
            total_descuentos += descuento
            rows_descuentos.append(['', grupo_descuento.upper(), self.add_coma( descuento )])
            kwargs_fields.setdefault('69cacdc9dcfb7b9636860e1d', []).append({
                '69cace21a461e2086b34e260': grupo_descuento.upper(),
                '69cace21a461e2086b34e261': round(descuento, 2)
            })
        
        rows_descuentos.append(['', 'TOTAL DESCUENTO', self.add_coma( total_descuentos )])
        # rows_descuentos.append([])
        # rows_descuentos.append([])
        # rows_descuentos.extend(rows_tecnicos)
        return total_descuentos, rows_descuentos, rows_tecnicos

    def genera_links(self, list_data_ocs):
        list_links = []
        for str_oc in list_data_ocs:
            list_links.append( f"LINK: {str_oc}" )
        return list_links

    def send_xls_email(self, nombre, email, id_conexion, xls, **kwargs):
        metadata = self.lkf_api.get_metadata(self.form_id_email)
        metadata['answers'] = {
            '69c98863121f86532adeb48d': self.fecha_periodo,
            '69c993457ef2fb39f071f3ba': nombre,
            '69c98863121f86532adeb48e': email,
            # '69c98863121f86532adeb48e': 'pluna@operacionpci.com.mx', # Test
            '69c98863121f86532adeb48f': id_conexion
        }
        metadata['answers'].update(xls)
        metadata['answers'].update(kwargs)
        resp_create_record_email = self.lkf_api.post_forms_answers(metadata)
        print(' - resp_create_record_email =',resp_create_record_email)

    def generar_excel_ocs(self):
        print(f'... ... Ejecutando proceso de Generar Excel de OCS CORTE = {self.fecha_periodo} ... ...')
        
        # Obtener los registros de Orden de Compra
        records_oc = lkf_obj.cr.aggregate( self.get_query_ocs() )

        count = 0 # Este contador nomas para mis pruebas
        group_OCs, group_Descuentos, group_Totales = {}, {}, {}
        for record in records_oc:
            count += 1
            # if count == 10:
            #     break
            
            contratista = record['connection_name']

            # Empiezo por la agrupacion
            group_OCs.setdefault( contratista, {} )
            group_Descuentos.setdefault( contratista, {'id': record['connection_id'], 'descuentos': set()} )
            group_Totales.setdefault( contratista, {'id': record['connection_id'], 'email': record['connection_email'], 'totales': set()} )
            for form in record['forms']:
                default_type_oc = 'ftth' if form['form_id'] in self.forms_ocs_ftth else 'cobre'

                # Se procesa cada folio dentro de la OC para identificar su agrupación y valores
                for folio_detail in form['detail_folios']:
                    type_folio = self.get_type_folio( folio_detail )
                    if not type_folio:
                        type_folio = default_type_oc
                    
                    group_OCs[ contratista ].setdefault(type_folio, []).append({
                        'importe': folio_detail.get('importe_a_pagar', 0),
                        'bono': folio_detail.get('bono_productividad', 0),
                        'incentivo_psr': folio_detail.get('incentivo_psr', 0),
                        'reparacion_instalaciones': folio_detail.get('reparacion_instalaciones', 0),
                        'folio_oc': folio_detail.get('folio_oc'),
                        'record_id_oc': folio_detail.get('record_id_oc'),
                        'retencion_resico': folio_detail.get('retencion_resico'),
                        'iva': folio_detail.get('iva'),
                        'total': folio_detail.get('total'),
                    })
                    
                    group_Descuentos[ contratista ]['descuentos'].add( self.get_str_descuentos(folio_detail) )
                    group_Totales[ contratista ]['totales'].add( self.get_str_totales(folio_detail) )
        
        # print(group_Descuentos)
        # print(group_Totales)

        for name_contratista, grupos_folios in group_OCs.items():
            print(f'\n === === === {name_contratista} === === ===')
            rows_xls, kwargs_fields = [], {}
            
            data_totales = group_Totales.get(name_contratista, {})
            email_contratista = data_totales.get('email')
            if not email_contratista:
                print('[ERROR] No hay email del contratista')
                continue

            rows_xls.append( [ name_contratista.upper() ] )
            rows_xls.append( [ '', 'CTD', 'TOTAL A FACTURAR' ] )
            
            total_monto_bonos, total_cantidad_bonos, subtotal_sin_descuentos = 0, 0, 0
            ordenes_de_compra = set()
            for grupo in ['a4', 'cobre', 'ftth', 'migraciones', 'quejas_psr', 'incentivo_psr']:
                list_folios = grupos_folios.get(grupo, [])
                total_facturar_grupo, monto_bonos, cant_bonos, set_ocs = self.process_importes(list_folios)
                total_monto_bonos += monto_bonos
                total_cantidad_bonos += cant_bonos
                ordenes_de_compra |= set_ocs
                name_group = grupo.replace('_', ' ').upper()
                cantidad_folios = len(list_folios)
                
                print(f'    - {name_group} | Cantidad = {cantidad_folios} | Total Facturar = {self.add_coma(total_facturar_grupo)}')
                print(f'        - {ordenes_de_compra}')

                row_xls = [ name_group, cantidad_folios, self.add_coma(total_facturar_grupo) ]

                kwargs_fields.setdefault('69cacc4d00615f9fe4ab283a', []).append({
                    '69cacccab6b64982b5ae927e': name_group,
                    '69cacccab6b64982b5ae927f': cantidad_folios,
                    '69cacccab6b64982b5ae9280': round(total_facturar_grupo, 2)
                })

                # Se generan los Links a las OCs
                cells_links = self.genera_links( set_ocs )
                row_xls.extend(cells_links)

                
                rows_xls.append( row_xls )
                subtotal_sin_descuentos += total_facturar_grupo
            
            print(f'     - BONO PRODUCTIVIDAD | Cantidad = {total_cantidad_bonos} | Total Facturar = {self.add_coma(total_monto_bonos)}')
            rows_xls.append( [ 'BONO PRODUCTIVIDAD', total_cantidad_bonos, self.add_coma(total_monto_bonos) ] )
            subtotal_sin_descuentos += total_monto_bonos
            rows_xls.append( [ 'SUBTOTAL SIN DESCUENTOS', '', self.add_coma(subtotal_sin_descuentos) ] )
            rows_xls.append([])
            rows_xls.append([])

            kwargs_fields['69cacdc9dcfb7b9636860e1c'] = round(subtotal_sin_descuentos, 2)

            # Se obtienen los descuentos
            monto_descuentos, tabla_descuentos, tabla_tecnicos = self.process_descuentos( group_Descuentos.get(name_contratista, {}), kwargs_fields )

            # TOTALES
            map_totales, folios_ocs_totales = self.parse_descuentos( data_totales.get('totales') )
            # print('map_totales =',dict(map_totales))

            # Matematica de totales
            subtotal = subtotal_sin_descuentos - monto_descuentos
            iva = subtotal * 0.16 # cuidado que hay contratistas donde aplica el 8%
            total = subtotal + iva

            rows_xls.extend( tabla_descuentos )
            rows_xls.append([])
            rows_xls.append(['APLICA AHORRO IMSS / INFONAVIT', '', 0])
            rows_xls.append([])
            rows_xls.append(['TOTAL A FACTURAR SIN IVA', '', self.add_coma( subtotal )])
            rows_xls.append(['IVA', '', self.add_coma( map_totales.get('iva', 0) )])
            rows_xls.append(['RESICO', '', self.add_coma( map_totales.get('retencion_resico', 0) )])
            rows_xls.append(['Total con IVA', '', self.add_coma( map_totales.get('total', 0) )])
            rows_xls.append([])
            rows_xls.extend( tabla_tecnicos )

            kwargs_fields['69cace87bba91641eb8c3f6a'] = round(monto_descuentos, 2)

            kwargs_fields['69cace87bba91641eb8c3f6b'] = round(subtotal, 2)
            kwargs_fields['69cace87bba91641eb8c3f6c'] = round(self.str_to_float(map_totales.get('iva', 0)), 2)
            kwargs_fields['69cace87bba91641eb8c3f6d'] = round(self.str_to_float(map_totales.get('retencion_resico', 0)), 2)
            kwargs_fields['69cace87bba91641eb8c3f6e'] = round(self.str_to_float(map_totales.get('total', 0)), 2)

            excel_ocs = self.generar_xls(rows_xls)
            if not excel_ocs:
                continue
            
            self.send_xls_email(name_contratista.upper(), email_contratista, data_totales.get('id'), excel_ocs, **kwargs_fields)

if __name__ == '__main__':
    lkf_obj = GenerarExcelOcs(settings, sys_argv=sys.argv)
    lkf_obj.console_run()
    lkf_obj.generar_excel_ocs()